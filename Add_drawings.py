#!/usr/bin/env python

from time import sleep
from turtle import clear
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
#from webdriver_manager.chrome import ChromeDriverManager
import keyboard
import openpyxl
from openpyxl.utils import get_column_letter

#Chrome driver
ser = Service(r"C:/chromedriver.exe")

#Chrome - Debugger Mode
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
driver = webdriver.Chrome(service=ser, options=options)

#Just for a reference - see console
print(driver.title)

#Lists to store the Excel values
element = []
sh_size = []
st_date = []
dwg_name = []
dwg_desc = []

#User Input
she_name = input("Please enter the Excel Sheet Name: ")
no_of_rows = input("Please enter the ending row number + 1: ")
no_of_rows = int(no_of_rows)


print(no_of_rows)

#Load workbook and sheet
file_location = "C:/Users/Arun/Documents/Timesheet/ESMT6.xlsx"
workbook = openpyxl.load_workbook(file_location, data_only=True)
sheet = workbook[she_name]

#Loops to store Excel values in the lists
col = 1
for row in range(3, no_of_rows):
    char = get_column_letter(col)
    cell_name = char + str(row)
    element.append(sheet[cell_name].value)

col=col+1
for row in range(3, no_of_rows):
    char = get_column_letter(col)
    cell_name = char + str(row)
    sh_size.append(sheet[cell_name].value)

col=col+1
for row in range(3, no_of_rows):
    char = get_column_letter(col)
    cell_name = char + str(row)
    dt = sheet[cell_name].value
    c_dt = dt.strftime('%d-%m-%Y')
    st_date.append(c_dt)

col=col+1
for row in range(3, no_of_rows):
    char = get_column_letter(col)
    cell_name = char + str(row)
    dwg_name.append(sheet[cell_name].value)

col=col+1
for row in range(3, no_of_rows):
    char = get_column_letter(col)
    cell_name = char + str(row)
    dwg_desc.append(sheet[cell_name].value)


#Printing just for a reference
print (element)
print (sh_size)
print (st_date)
print (dwg_name)
print (dwg_desc)


sleep(1)

#Clicks and Enters the lists' values into the form in a loop
for i in range(len(element)):

    #Click - New button
    elem2 = driver.find_element("xpath", '/html/body/app-root/app-project-detail/div/div[4]/div/button')
    elem2.click()
    sleep(1)
    
    print("\n")
    print(" | ")

    #Fields

    #Element
    print (element[i], " | ")
    elem3 = driver.find_element("xpath", '/html/body/app-root/app-project-detail/div/div[4]/div/div/div/div/div[1]/ng-autocomplete/div[1]/div[1]/input')
    elem3.clear()
    elem3.send_keys(element[i])

    #Sheet Size - Dropdown
    print (sh_size[i], " | ")
    sleep(0.5)
    elem5 = Select(driver.find_element("xpath", '/html/body/app-root/app-project-detail/div/div[4]/div/div/div/div/select')).select_by_visible_text(sh_size[i])
    sleep(0.5)

    #Start Date
    print (st_date[i], " | ")
    elem5 = driver.find_element("xpath", '/html/body/app-root/app-project-detail/div/div[4]/div/div/div/div/input[1]')
    elem5.clear()
    elem5.send_keys(st_date[i])
    sleep(0.5)

    #Drawing name/number
    print (dwg_name[i], " | ")
    elem6 = driver.find_element("xpath", '/html/body/app-root/app-project-detail/div/div[4]/div/div/div/div/input[2]')
    elem6.clear()
    elem6.send_keys(dwg_name[i])

    #Drawing Desc.
    print (dwg_desc[i], " | ")
    elem7 = driver.find_element("xpath", '/html/body/app-root/app-project-detail/div/div[4]/div/div/div/div/input[3]')
    elem7.clear()
    elem7.send_keys(dwg_desc[i])
    
    #Hotkey to wait before proceeding
    keyboard.wait("esc")

    

