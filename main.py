from time import sleep
from turtle import clear
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import keyboard
import openpyxl
from openpyxl.utils import get_column_letter

chrome_driver = "C:/chromedriver.exe"

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)
print(driver.title)

elem = driver.find_element("xpath", '/html/body/app-root/app-landing/div/div/app-login/div/div[1]/input')
elem.send_keys("arun.venkatesh@csiglobal.net")

elem = driver.find_element("xpath", '/html/body/app-root/app-landing/div/div/app-login/div/div[2]/input')
elem.send_keys("projectinc0.")

sleep(0.5)

elem = driver.find_element("xpath", '/html/body/app-root/app-landing/div/div/app-login/div/div[3]/input')
elem.click()

sleep(5)

driver.get("https://projectmanagement-8d5c4.web.app/userManager")

name = []
short_form = []
e_Address = []
passwd = []


file_location = "C:/Users/Arun/Documents/Timesheet/Timesheet_Team_Copy.xlsx"
workbook = openpyxl.load_workbook(file_location)
sheet = workbook["Team7"]
col = 2
for row in range(3, 11):
    char = get_column_letter(col)
    cell_name = char + str(row)
    name.append(sheet[cell_name].value)
    
col=col+1
for row in range(3, 11):
    char = get_column_letter(col)
    cell_name = char + str(row)
    short_form.append(sheet[cell_name].value)

col=col+1
for row in range(3, 11):
    char = get_column_letter(col)
    cell_name = char + str(row)
    e_Address.append(sheet[cell_name].value)

col=col+1
for row in range(3, 11):
    char = get_column_letter(col)
    cell_name = char + str(row)
    passwd.append(sheet[cell_name].value)

print (name)
print (short_form)
print (e_Address)
print (passwd)

sleep(1)

for i in range(len(name)):
    elem2 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div/div[1]/button')
    elem2.click()
    sleep(1)
    elem3 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[1]/input')
    elem3.send_keys(name[i])
    elem4 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[2]/input')
    elem4.send_keys(short_form[i])
    elem5 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[3]/input')
    elem5.send_keys(e_Address[i])
    elem6 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[4]/select')
    elem6.click
    sleep(0.5)
    elem7 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[4]/select/option[3]')
    elem7.click()
    sleep(2)
    elem8 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[5]/input')
    elem8.send_keys(passwd[i])
    sleep(0.5)
    elem8 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[6]/input')
    elem8.send_keys(passwd[i])
    sleep(1)
    keyboard.wait("esc")

