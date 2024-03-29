from time import sleep
from turtle import clear
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import keyboard
import openpyxl
from openpyxl.utils import get_column_letter

#Chrome driver
chrome_driver = "C:/chromedriver.exe"

#Chrome - Debugger Mode
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)

#Just for a reference
print(driver.title)

#Lists to store the Excel values
name = []
short_form = []
e_Address = []
passwd = []

#Load workbook and sheet
file_location = "C:/Users/Arun/Documents/Timesheet/Timesheet_Team_Copy.xlsx"
workbook = openpyxl.load_workbook(file_location)
sheet = workbook["Team7"]

#Loops to store values in the lists
col = 2
for row in range(3, 11):
    char = get_column_letter(col)
    cell_name = char + str(row)
    name.append(sheet[cell_name].value)
    
col=col+1
for row in range(3, 11):
    char = get_column_letter(col)
    cell_name = char + str(row)
    short_form.append(sheet[cell_name].value)

col=col+1
for row in range(3, 11):
    char = get_column_letter(col)
    cell_name = char + str(row)
    e_Address.append(sheet[cell_name].value)

col=col+1
for row in range(3, 11):
    char = get_column_letter(col)
    cell_name = char + str(row)
    passwd.append(sheet[cell_name].value)

#Printing just for a reference
print (name)
print (short_form)
print (e_Address)
print (passwd)

sleep(1)


for i in range(len(name)):
    #Click - New User button
    elem2 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div/div[1]/button')
    elem2.click()
    sleep(1)
    #Fields
    elem3 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[1]/input')
    elem3.send_keys(name[i])
    elem4 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[2]/input')
    elem4.send_keys(short_form[i])
    elem5 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[3]/input')
    elem5.send_keys(e_Address[i])
    #Dropdown
    elem6 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[4]/select')
    elem6.click
    sleep(0.5)
    elem7 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[4]/select/option[3]')
    elem7.click()
    sleep(2)
    elem8 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[5]/input')
    elem8.send_keys(passwd[i])
    sleep(0.5)
    elem8 = driver.find_element("xpath", '/html/body/app-root/app-user-manager/div[2]/div/div/div/div[6]/input')
    elem8.send_keys(passwd[i])
    sleep(1)
    #Hotkey to wait before the next iteration
    keyboard.wait("esc")

